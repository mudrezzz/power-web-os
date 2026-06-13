from __future__ import annotations

from pathlib import Path
from typing import Any
import warnings

from openpyxl import load_workbook

from power_web_os.icp_radar import (
    AccountQualificationModel,
    AtomicRule,
    CRITERION_CODES,
    GlobalSearchPolicy,
    ICPProfile,
    ICPRadar,
    ICPRadarArtifact,
    ICPRadarCandidate,
    IntentSignalDefinition,
    MonitoringPolicy,
    RadarDefinition,
    RadarDefinitionValidator,
    RadarMetadata,
    RadarScoringModel,
    RadarValidationReport,
    RuleGroup,
    SignalScoreRule,
    SignalCriterion,
    SignalScoringRubric,
    SourceDefinition,
    EvidenceSource,
    SourcePolicy,
)
from power_web_os.icp_radar_evidence import CriterionEvidenceBuilder, load_criterion_evidence_fixture

warnings.filterwarnings("ignore", category=UserWarning, message=".*extension is not supported.*")

REQUIRED_SHEETS = ("Summary", "ICP Matrix", "Criteria", "Sources")


class ICPRadarWorkbookError(ValueError):
    pass


def load_icp_radar_workbook(path: Path) -> ICPRadarArtifact:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(path, data_only=True, read_only=True)
        return _artifact_from_workbook(workbook, path)


def _artifact_from_workbook(workbook: Any, path: Path) -> ICPRadarArtifact:
    missing_sheets = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing_sheets:
        raise ICPRadarWorkbookError(f"ICP Radar workbook misses sheets: {', '.join(missing_sheets)}")

    criteria = _read_criteria(workbook["Criteria"])
    sources = _read_sources(workbook["Sources"])
    summary_overlay = _read_summary_overlay(workbook["Summary"])
    source_refs_by_url = {source.url: source.source_id for source in sources if source.url}
    evidence_fixture_path = path.parent / "toir_sibur_criterion_evidence.json"
    evidence_fixture = load_criterion_evidence_fixture(evidence_fixture_path)
    radar = ICPRadar()
    candidates = []

    for row_index, row in enumerate(_value_rows(workbook["ICP Matrix"], min_row=2), start=1):
        legal_name = _text(_get(row, 2))
        if not legal_name:
            continue

        number = _text(_get(row, 0)) or str(row_index)
        criteria_scores = {
            code: _to_int(_get(row, 16 + offset))
            for offset, code in enumerate(CRITERION_CODES)
        }
        score = radar.build_score(criteria_scores)
        source_urls = tuple(_split_lines(_text(_get(row, 10))))
        evidence_refs = tuple(source_refs_by_url.get(url, url) for url in source_urls)
        overlay = summary_overlay.get(number) or summary_overlay.get(legal_name) or {}

        candidates.append(
            ICPRadarCandidate(
                rank=0,
                account_id=f"icp-sibur-{_stable_number(number):03d}",
                ppo=_text(_get(row, 1)),
                legal_name=legal_name,
                account_type=_text(_get(row, 3)),
                description=_text(_get(row, 4)),
                inn=_text(_get(row, 5)),
                revenue=_text(_get(row, 6)),
                site=_text(_get(row, 7)),
                confidence=_text(_get(row, 8)),
                signal_summary=_text(_get(row, 9)),
                main_signal=str(overlay.get("main_signal") or _text(_get(row, 9))),
                comment=str(overlay.get("comment") or ""),
                source_urls=source_urls,
                evidence_refs=evidence_refs,
                criteria_scores=criteria_scores,
                criteria_evidence={},
                score=score,
            )
        )

    ranked_candidates = radar.rank(candidates)
    ranked_candidates = CriterionEvidenceBuilder(
        criteria=tuple(criteria),
        sources=tuple(sources),
        fixture=evidence_fixture,
    ).attach(ranked_candidates)
    return ICPRadarArtifact(
        profile=_profile(path),
        criteria=tuple(criteria),
        sources=tuple(sources),
        definition=_definition(
            path=path,
            criteria=tuple(criteria),
            sources=tuple(sources),
        ),
        candidates=ranked_candidates,
        workflow_metadata={
            "workflow_name": "ICPRadarXlsxImport",
            "artifact_version": "0.6.5.2",
            "source_workbook": path.name,
            "criteria_evidence_contract_version": "0.6.2.3",
            "criteria_evidence_fixture": evidence_fixture_path.name,
            "sheet_names": list(workbook.sheetnames),
            "candidate_count": len(ranked_candidates),
            "criteria_count": len(criteria),
            "source_count": len(sources),
            "scoring": "workbook-compatible deterministic formula",
        },
    )


def _definition(
    *,
    path: Path,
    criteria: tuple[SignalCriterion, ...],
    sources: tuple[EvidenceSource, ...],
) -> RadarDefinition:
    source_definitions = _source_definitions_from_workbook(sources)
    source_ids = tuple(source.source_id for source in source_definitions)
    definition = RadarDefinition(
        definition_id="radar-def-toir-sibur",
        metadata=RadarMetadata(
            name="ТОиР / SIBUR",
            description="ICP Radar для поиска юридических лиц СИБУР с релевантностью к автоматизации ТОиР и промышленной аналитике.",
            owner="ABM Research",
            status="active",
        ),
        global_search_policy=GlobalSearchPolicy(
            sources=source_definitions,
            keywords=("СИБУР ТОиР", "СИБУР EAM", "СИБУР предиктивная аналитика", "СИБУР модернизация ремонтов"),
            exclusions=(
                "Непрофильные сервисные юрлица без производственного контура",
                "Дочерние структуры без открытых источников по ТОиР",
            ),
            allow_system_sources=True,
        ),
        account_qualification=AccountQualificationModel(
            rule_group=RuleGroup(
                group_id="qualification-root",
                name="Account qualification rules",
                operator="AND",
                rules=(
                    AtomicRule(
                        rule_id="qualification-sibur-group",
                        name="SIBUR group membership",
                        description="Компания входит в группу СИБУР.",
                        target_field="holding",
                        comparison_operator="equals",
                        value="СИБУР",
                        requirement_level="required",
                        source_policy=SourcePolicy(
                            source_ids=("sibur_workbook",),
                            source_logic="OR",
                            allow_additional_sources=True,
                            fallback_confidence="medium",
                        ),
                    ),
                    AtomicRule(
                        rule_id="qualification-industry",
                        name="Industrial profile",
                        description="Компания относится к нефтехимии, нефтегазу или промышленным производственным активам.",
                        target_field="industry",
                        comparison_operator="in",
                        value="нефтехимия, нефтегаз, промышленное производство",
                        requirement_level="required",
                        source_policy=SourcePolicy(
                            source_ids=("sibur_workbook",),
                            source_logic="OR",
                            allow_additional_sources=True,
                            fallback_confidence="medium",
                        ),
                    ),
                    AtomicRule(
                        rule_id="qualification-revenue",
                        name="Enterprise scale",
                        description="Выручка выше 10 млрд рублей или есть сопоставимый enterprise-масштаб.",
                        target_field="revenue_billion_rub",
                        comparison_operator="greater_than",
                        value="10",
                        requirement_level="recommended",
                        source_policy=SourcePolicy(
                            source_ids=("sbis",),
                            source_logic="OR",
                            allow_additional_sources=True,
                            fallback_confidence="low",
                        ),
                    ),
                ),
            )
        ),
        intent_signals=tuple(_intent_signal_from_criterion(criterion, source_ids) for criterion in criteria),
        monitoring_policy=MonitoringPolicy(
            cadence="monthly",
            lookback_window="90 days",
            run_mode="incremental_signal_monitoring",
            deduplication="ignore_evidence_refs_seen_in_previous_runs",
            stale_after="180 days",
        ),
        scoring_model=RadarScoringModel(
            fit_model={
                "formula_preset": "weighted_average",
                "description": "Account fit is calculated from account qualification rules.",
                "custom_formula": "",
                "uses": ["qualification-sibur-group", "qualification-industry", "qualification-revenue"],
            },
            intent_model={
                "formula_preset": "weighted_average",
                "description": "Intent is calculated from observed interest signals C1-C20.",
                "custom_formula": "",
                "uses": list(CRITERION_CODES),
            },
            tier_model={
                "basis": "fit + intent",
                "description": "Tier classifies candidates using fit and intent thresholds.",
            },
            tier_thresholds={
                "Tier 1": ">=38",
                "Tier 2": ">=25",
                "Tier 3": ">=15",
                "Monitor": "<15",
            },
            confidence_penalties={
                "low": "-20%",
                "medium": "-10%",
                "high": "0%",
                "none": "exclude from positive score",
            },
        ),
        validation_report=RadarValidationReport(errors=(), warnings=(), info=()),
    )
    return _definition_with_validation(definition)


def _source_definitions_from_workbook(sources: tuple[EvidenceSource, ...]) -> tuple[SourceDefinition, ...]:
    workbook_sources = tuple(
        SourceDefinition(
            source_id=source.source_id,
            source_type="url",
            label=source.usage or source.source_id,
            reference=source.url,
            trust_level="medium",
        )
        for source in sources
    )
    return (
        SourceDefinition(
            source_id="sibur_workbook",
            source_type="manual_dataset",
            label="SIBUR ICP workbook",
            reference="demo/fixtures/icp_radar/sibur_icp_pass1.xlsx",
            trust_level="high",
        ),
        SourceDefinition(
            source_id="sbis",
            source_type="api",
            label="СБИС",
            reference="https://sbis.ru",
            trust_level="high",
        ),
        SourceDefinition(
            source_id="yandex_search",
            source_type="search_engine",
            label="Яндекс поиск",
            reference="https://ya.ru",
            trust_level="medium",
        ),
        *workbook_sources,
    )


def _intent_signal_from_criterion(
    criterion: SignalCriterion,
    source_ids: tuple[str, ...],
) -> IntentSignalDefinition:
    source_policy = SourcePolicy(
        source_ids=source_ids[: min(3, len(source_ids))],
        source_logic="OR",
        allow_additional_sources=True,
        fallback_confidence="low",
    )
    return IntentSignalDefinition(
        signal_id=f"signal-{criterion.code.lower()}",
        code=criterion.code,
        name=criterion.name,
        description=criterion.description,
        trigger_rule_group=RuleGroup(
            group_id=f"trigger-{criterion.code.lower()}",
            operator="OR",
            rules=(
                AtomicRule(
                    rule_id=f"trigger-{criterion.code.lower()}-mention",
                    description=criterion.description,
                    target_field="public_evidence",
                    comparison_operator="contains",
                    value=criterion.name,
                    requirement_level="recommended",
                    source_policy=source_policy,
                ),
            ),
        ),
        source_policy=source_policy,
        scoring_rubric=SignalScoringRubric(
            scale=(0, 1, 2),
            rules=(
                SignalScoreRule(
                    score=0,
                    description="Сигнал не наблюдается или не подтвержден источниками.",
                    rule_group=RuleGroup(
                        group_id=f"rubric-{criterion.code.lower()}-0",
                        operator="AND",
                        rules=(
                            AtomicRule(
                                rule_id=f"rubric-{criterion.code.lower()}-0-rule",
                                description="Нет подтвержденного наблюдения по сигналу.",
                                target_field=criterion.code,
                                comparison_operator="equals",
                                value="0",
                                requirement_level="recommended",
                                source_policy=source_policy,
                            ),
                        ),
                    ),
                ),
                SignalScoreRule(
                    score=1,
                    description="Есть косвенное или слабое наблюдение, требующее проверки.",
                    rule_group=RuleGroup(
                        group_id=f"rubric-{criterion.code.lower()}-1",
                        operator="AND",
                        rules=(
                            AtomicRule(
                                rule_id=f"rubric-{criterion.code.lower()}-1-rule",
                                description="Есть один слабый источник или непрямой индикатор.",
                                target_field=criterion.code,
                                comparison_operator="equals",
                                value="1",
                                requirement_level="recommended",
                                source_policy=source_policy,
                            ),
                        ),
                    ),
                ),
                SignalScoreRule(
                    score=2,
                    description="Сигнал подтвержден релевантным источником и может усиливать скоринг.",
                    rule_group=RuleGroup(
                        group_id=f"rubric-{criterion.code.lower()}-2",
                        operator="AND",
                        rules=(
                            AtomicRule(
                                rule_id=f"rubric-{criterion.code.lower()}-2-rule",
                                description="Есть сильное наблюдение или несколько согласованных источников.",
                                target_field=criterion.code,
                                comparison_operator="equals",
                                value="2",
                                requirement_level="recommended",
                                source_policy=source_policy,
                            ),
                        ),
                    ),
                ),
            ),
        ),
    )


def _definition_with_validation(definition: RadarDefinition) -> RadarDefinition:
    report = RadarDefinitionValidator().validate(definition)
    return RadarDefinition(
        definition_id=definition.definition_id,
        metadata=definition.metadata,
        global_search_policy=definition.global_search_policy,
        account_qualification=definition.account_qualification,
        intent_signals=definition.intent_signals,
        monitoring_policy=definition.monitoring_policy,
        scoring_model=definition.scoring_model,
        validation_report=report,
    )


def _profile(path: Path) -> ICPProfile:
    return ICPProfile(
        profile_id="toir-sibur",
        name="ТОиР automation ICP Radar",
        product="Автоматизация ТОиР",
        holding="СИБУР",
        run_mode="fixture_import",
        source_workbook=path.name,
        scoring_formula={
            "fit_score": "C13 + C14 + C15 + C16 + C17",
            "intent_score": "C1..C9 + C18 + C19",
            "trigger_score": "C10 + C11 + C12 + C20",
            "total_score": "sum(C1..C20)",
            "tiers": {
                "Tier 1": ">=38",
                "Tier 2": ">=25",
                "Tier 3": ">=15",
                "Monitor": "<15",
            },
        },
    )


def _read_criteria(sheet: Any) -> list[SignalCriterion]:
    criteria = []
    for row in _value_rows(sheet, min_row=2):
        code = _text(_get(row, 0))
        if not code:
            continue
        criteria.append(
            SignalCriterion(
                code=code,
                name=_text(_get(row, 1)),
                description=_text(_get(row, 2)),
                scoring_guidance=_text(_get(row, 3)),
            )
        )
    return criteria


def _read_sources(sheet: Any) -> list[EvidenceSource]:
    sources = []
    for row_index, row in enumerate(_value_rows(sheet, min_row=2), start=1):
        source_id = _text(_get(row, 0)) or f"S{row_index}"
        url = _text(_get(row, 1))
        if not url:
            continue
        sources.append(
            EvidenceSource(
                source_id=source_id,
                url=url,
                usage=_text(_get(row, 2)),
            )
        )
    return sources


def _read_summary_overlay(sheet: Any) -> dict[str, dict[str, str]]:
    overlay: dict[str, dict[str, str]] = {}
    for row in _value_rows(sheet, min_row=2):
        number = _text(_get(row, 1))
        legal_name = _text(_get(row, 2))
        data = {
            "main_signal": _text(_get(row, 6)),
            "comment": _text(_get(row, 7)),
        }
        if number:
            overlay[number] = data
        if legal_name:
            overlay[legal_name] = data
    return overlay


def _value_rows(sheet: Any, *, min_row: int) -> Any:
    return sheet.iter_rows(min_row=min_row, values_only=True)


def _get(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    return int(float(value))


def _split_lines(value: str) -> list[str]:
    return [item.strip() for item in value.replace(";", "\n").splitlines() if item.strip()]


def _stable_number(value: str) -> int:
    try:
        return int(float(value))
    except ValueError:
        return sum((index + 1) * ord(char) for index, char in enumerate(value)) % 100000
