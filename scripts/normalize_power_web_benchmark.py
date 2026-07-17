"""Normalize the approved public subset of the SIBUR contacts workbook.

The workbook is an intake source, not a repository artifact. This script reads
only public role/source columns and never copies contact or outreach fields.
"""

from __future__ import annotations

import argparse
from datetime import UTC, date, datetime
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from power_web_os.application.radar.power_web_discovery.benchmark import (
    EmploymentControl,
    IdentityPairControl,
    PowerWebBenchmark,
    PowerWebBenchmarkFreeze,
    PowerWebBenchmarkPlanningContext,
    PowerWebBlindControls,
    ProfileControl,
    RelationshipControl,
    benchmark_sha256,
)
from power_web_os.application.radar.power_web_discovery.contracts import RoleDemand


AS_OF = date(2026, 7, 17)
BENCHMARK_ID = "sibur-priority-power-web"
BENCHMARK_VERSION = "1.0.0"


def _profile(
    *,
    profile_ref: str,
    name: str | None,
    employer: str,
    title: str,
    source_lane: str,
    url: str,
    role_ids: tuple[str, ...],
    facts: tuple[str, ...],
) -> ProfileControl:
    return ProfileControl(
        control_id=f"profile-control-{profile_ref}",
        profile_ref=profile_ref,
        anonymous=name is None,
        expected_display_name=name,
        expected_employer=employer,
        expected_title=title,
        source_lane=source_lane,
        expected_public_facts=facts,
        provenance_urls=(url,),
        expected_role_demand_ids=role_ids,
        as_of=AS_OF,
        expected_state="retained_profile",
    )


def _role_policy() -> tuple[RoleDemand, ...]:
    rows = (
        ("role-chief-engineer", "Главный инженер / технический руководитель", ("chief engineer", "технический директор")),
        ("role-production-director", "Директор по производству", ("руководитель производства", "production director")),
        ("role-maintenance-leader", "Руководитель ТОиР и ремонтов", ("начальник ТОиР", "руководитель ремонтов")),
        ("role-reliability-leader", "Руководитель надежности", ("главный эксперт по надежности", "reliability leader")),
        ("role-chief-technologist", "Главный технолог", ("руководитель технологии", "chief technologist")),
        ("role-energy-director", "Руководитель энергообеспечения", ("главный энергетик", "energy director")),
        ("role-automation-leader", "Руководитель АСУТП / КИПиА", ("руководитель автоматизации", "automation leader")),
        ("role-digital-transformation-leader", "Руководитель цифровизации производства", ("Industry 4.0", "предиктивная диагностика")),
    )
    return tuple(
        RoleDemand(
            demand_id=demand_id,
            role=role,
            scope="account_and_production_site",
            aliases=aliases,
            expected_evidence=("public employer", "public role title", "source date or current-page basis"),
            reason="Роль влияет на надежность, ТОиР, эксплуатацию или внедрение SmartDiagnostics.",
        )
        for demand_id, role, aliases in rows
    )


def _benchmark() -> PowerWebBenchmark:
    zhalyuk_official = "https://www.sibur.ru/zapsibneftekhim/press-center/sotrudniki-zapsibneftekhima-proveli-proforientatsionnyy-urok-dlya-shkolnikov-tobolska-/"
    zhalyuk_industry = "https://eepir.ru/new/cifrovaya-transformaciya-rza-tekushhee-sostoyanie-problematika-perspektivy/"
    malyavin_official = "https://www.sibur.ru/rusvinyl/press-center/rusvinil-podderzhal-provedenie-mezhdunarodnogo-ekologicheskogo-telefestivalya/"
    malyavin_industry = "https://vz-nn.ru/news/promyshlennost/72921/"
    shein_official = "https://www.sibur.ru/SiburKstovo/press-center/benzol-i-etilen-proizvodstva-sibur-kstovo-poluchili-nagrady-za-kachestvo/"
    shein_industry = "https://vz-nn.ru/news/promyshlennost/73096/"
    stroev_official = "https://www.sibur.ru/kazanorgsintez/ru/press-center/sibur-rasshiryaet-proekt-po-provedeniyu-interaktivnykh-urokov-khimii-mendeleevskaya-smena/"
    stroev_publication = "https://kazanfirst.ru/news/bolee-47-tysyachi-shkolnikov-i-studentov-tatarstana-stali-uchastnikami-mendeleevskoj-smeny"
    muzafarov_publication = "https://www.mrc.ru/news/417832-evgeniy-gorobec-naznachen-generalnim-direktorom-npp-neftehimiya"
    anonymous_hh = "https://hh.ru/resume/83f66d36000c499aec0027df83436b4b726665"

    profiles = (
        _profile(
            profile_ref="zhalyuk-official-2026",
            name="Роман Жалюк",
            employer="ООО «ЗапСибНефтехим»",
            title="Главный инженер",
            source_lane="official_company",
            url=zhalyuk_official,
            role_ids=("role-chief-engineer", "role-digital-transformation-leader"),
            facts=("Роман Жалюк назван главным инженером предприятия.", "Упомянуты цифровые технологии, ИИ и прогнозирование."),
        ),
        _profile(
            profile_ref="zhalyuk-industry-2025",
            name="Роман Жалюк",
            employer="ООО «ЗапСибНефтехим»",
            title="Главный инженер",
            source_lane="industry_web",
            url=zhalyuk_industry,
            role_ids=("role-chief-engineer", "role-digital-transformation-leader"),
            facts=("Роман Жалюк назван главным инженером ЗапСибНефтехима.", "Материал посвящен цифровизации и надежности РЗА."),
        ),
        _profile(
            profile_ref="malyavin-official-2024",
            name="Алексей Малявин",
            employer="ООО «РусВинил»",
            title="Главный инженер",
            source_lane="official_company",
            url=malyavin_official,
            role_ids=("role-chief-engineer", "role-reliability-leader"),
            facts=("Алексей Малявин назван главным инженером РусВинила.",),
        ),
        _profile(
            profile_ref="malyavin-industry-2025",
            name="Алексей Малявин",
            employer="ООО «СИБУР-Кстово» / ООО «РусВинил»",
            title="Главный инженер",
            source_lane="industry_web",
            url=malyavin_industry,
            role_ids=("role-chief-engineer", "role-reliability-leader", "role-maintenance-leader"),
            facts=("Алексей Малявин назван главным инженером двух предприятий.", "Он описывает RBI и прогнозирование сроков ремонта."),
        ),
        _profile(
            profile_ref="shein-official-2024",
            name="Андрей Шеин",
            employer="ООО «СИБУР-Кстово» / ООО «РусВинил»",
            title="Главный технолог",
            source_lane="official_company",
            url=shein_official,
            role_ids=("role-chief-technologist",),
            facts=("Андрей Шеин назван главным технологом двух предприятий.",),
        ),
        _profile(
            profile_ref="shein-industry-2025",
            name="Андрей Шеин",
            employer="ООО «РусВинил»",
            title="Главный технолог",
            source_lane="industry_web",
            url=shein_industry,
            role_ids=("role-chief-technologist",),
            facts=("Андрей Шеин назван главным технологом РусВинила.",),
        ),
        _profile(
            profile_ref="stroev-official-2025",
            name="Анатолий Строев",
            employer="ПАО «Казаньоргсинтез»",
            title="Главный инженер",
            source_lane="official_company",
            url=stroev_official,
            role_ids=("role-chief-engineer",),
            facts=("Анатолий Строев назван главным инженером Казаньоргсинтеза.",),
        ),
        _profile(
            profile_ref="stroev-publication-2025",
            name="Анатолий Строев",
            employer="ПАО «Казаньоргсинтез»",
            title="Главный инженер",
            source_lane="publications_events",
            url=stroev_publication,
            role_ids=("role-chief-engineer",),
            facts=("Анатолий Строев назван главным инженером Казаньоргсинтеза.",),
        ),
        _profile(
            profile_ref="muzafarov-publication-2025",
            name="Рустем Музафаров",
            employer="ПАО «Казаньоргсинтез»",
            title="Директор по производству; ранее главный инженер",
            source_lane="industry_web",
            url=muzafarov_publication,
            role_ids=("role-production-director", "role-chief-engineer"),
            facts=("Назначен директором по производству.", "Ранее занимал должность главного инженера предприятия."),
        ),
        _profile(
            profile_ref="anonymous-zapsib-toir-hh",
            name=None,
            employer="ООО «ЗапСибНефтехим»",
            title="Начальник отдела ТОиР",
            source_lane="hh_public_web",
            url=anonymous_hh,
            role_ids=("role-maintenance-leader",),
            facts=("Публично индексируемое HH-резюме содержит работодателя и роль.", "ФИО и актуальность занятости не подтверждены."),
        ),
    )
    same_pairs = (
        ("zhalyuk", "zhalyuk-official-2026", "zhalyuk-industry-2025"),
        ("malyavin", "malyavin-official-2024", "malyavin-industry-2025"),
        ("shein", "shein-official-2024", "shein-industry-2025"),
        ("stroev", "stroev-official-2025", "stroev-publication-2025"),
    )
    different_pairs = (
        ("zhalyuk-vs-malyavin", "zhalyuk-official-2026", "malyavin-official-2024"),
        ("zhalyuk-vs-shein", "zhalyuk-official-2026", "shein-official-2024"),
        ("malyavin-vs-stroev", "malyavin-industry-2025", "stroev-official-2025"),
        ("shein-vs-stroev", "shein-industry-2025", "stroev-publication-2025"),
    )
    profile_by_ref = {item.profile_ref: item for item in profiles}
    identity_pairs = tuple(
        IdentityPairControl(
            control_id=f"identity-same-{name}",
            left_profile_ref=left,
            right_profile_ref=right,
            expected_state="probable",
            provenance_urls=profile_by_ref[left].provenance_urls + profile_by_ref[right].provenance_urls,
            as_of=AS_OF,
        )
        for name, left, right in same_pairs
    ) + tuple(
        IdentityPairControl(
            control_id=f"identity-different-{name}",
            left_profile_ref=left,
            right_profile_ref=right,
            expected_state="confirmed_different",
            provenance_urls=profile_by_ref[left].provenance_urls + profile_by_ref[right].provenance_urls,
            as_of=AS_OF,
        )
        for name, left, right in different_pairs
    )
    employment = (
        EmploymentControl(
            control_id="employment-zhalyuk-current",
            subject_ref="zhalyuk-official-2026",
            employer="ООО «ЗапСибНефтехим»",
            title="Главный инженер",
            expected_state="current",
            provenance_urls=(zhalyuk_official,),
            as_of=AS_OF,
        ),
        EmploymentControl(
            control_id="employment-muzafarov-former-chief-engineer",
            subject_ref="muzafarov-publication-2025",
            employer="ПАО «Казаньоргсинтез»",
            title="Главный инженер",
            expected_state="former",
            provenance_urls=(muzafarov_publication,),
            as_of=AS_OF,
        ),
        EmploymentControl(
            control_id="employment-anonymous-hh-unknown",
            subject_ref="anonymous-zapsib-toir-hh",
            employer="ООО «ЗапСибНефтехим»",
            title="Начальник отдела ТОиР",
            expected_state="unknown",
            provenance_urls=(anonymous_hh,),
            as_of=AS_OF,
        ),
    )
    relationships = (
        RelationshipControl(
            control_id="relationship-zhalyuk-chief-engineer",
            source_ref="zhalyuk-official-2026",
            target_ref="role-chief-engineer",
            relationship_type="occupies_role",
            expected_state="confirmed",
            provenance_urls=(zhalyuk_official,),
            as_of=AS_OF,
        ),
        RelationshipControl(
            control_id="relationship-stroev-chief-engineer",
            source_ref="stroev-official-2025",
            target_ref="role-chief-engineer",
            relationship_type="occupies_role",
            expected_state="confirmed",
            provenance_urls=(stroev_official,),
            as_of=AS_OF,
        ),
        RelationshipControl(
            control_id="relationship-malyavin-reliability-influence",
            source_ref="malyavin-industry-2025",
            target_ref="role-reliability-leader",
            relationship_type="influences_maintenance_and_reliability",
            expected_state="review_needed",
            provenance_urls=(malyavin_industry,),
            as_of=AS_OF,
        ),
        RelationshipControl(
            control_id="relationship-anonymous-maintenance-role",
            source_ref="anonymous-zapsib-toir-hh",
            target_ref="role-maintenance-leader",
            relationship_type="possible_role_occupancy",
            expected_state="review_needed",
            provenance_urls=(anonymous_hh,),
            as_of=AS_OF,
        ),
    )
    return PowerWebBenchmark(
        benchmark_id=BENCHMARK_ID,
        benchmark_version=BENCHMARK_VERSION,
        as_of=AS_OF,
        status="user_accepted",
        planning_context=PowerWebBenchmarkPlanningContext(
            account_id="sibur-priority-industrial-contour",
            account_name="СИБУР: приоритетный промышленный контур",
            product_context="SmartDiagnostics: предиктивная диагностика, надежность оборудования и ТОиР.",
            role_policy=_role_policy(),
            allowed_source_lanes=(
                "hh_public_web",
                "official_company",
                "professional_networks",
                "publications_events",
                "industry_web",
                "generic_web",
            ),
        ),
        blind_controls=PowerWebBlindControls(
            profiles=profiles,
            identity_pairs=identity_pairs,
            employment=employment,
            relationships=relationships,
        ),
    )


def _workbook_public_rows(source: Path) -> dict[str, dict[int, tuple[str, str, str, str]]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    summary = workbook.worksheets[0]
    contacts = workbook["Contacts"]
    return {
        summary.title: {
            row: (
                str(summary.cell(row, 2).value or ""),
                str(summary.cell(row, 6).value or ""),
                str(summary.cell(row, 7).value or ""),
                str(summary.cell(row, 15).value or ""),
            )
            for row in (3, 13, 27, 28, 46)
        },
        contacts.title: {
            56: (
                str(contacts.cell(56, 2).value or ""),
                str(contacts.cell(56, 3).value or ""),
                str(contacts.cell(56, 4).value or ""),
                str(contacts.cell(56, 12).value or ""),
            )
        },
    }


def _verify_source_rows(rows: dict[str, dict[int, tuple[str, str, str, str]]]) -> None:
    required = {
        3: ("Жалюк", "eepir.ru"),
        13: ("na", "hh.ru/resume/83f66d36"),
        27: ("Малявин", "vz-nn.ru/news/promyshlennost/72921"),
        28: ("Шеин", "vz-nn.ru/news/promyshlennost/73096"),
        46: ("Строев", "kazanfirst.ru"),
    }
    summary_rows = next(iter(rows.values()))
    for row_number, tokens in required.items():
        encoded = " | ".join(summary_rows[row_number])
        if not all(token.casefold() in encoded.casefold() for token in tokens):
            raise ValueError(f"source workbook row {row_number} no longer matches benchmark mapping")
    contacts_rows = rows["Contacts"]
    encoded = " | ".join(contacts_rows[56])
    if "Музафаров" not in encoded or "ранее главный инженер" not in encoded:
        raise ValueError("Contacts row 56 no longer proves the former-employment control")


def _source_record(*, source: Path, rows: dict[str, dict[int, tuple[str, str, str, str]]]) -> dict[str, Any]:
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    return {
        "schema_version": "power_web_benchmark_source.v1",
        "source_file_name": source.name,
        "source_sha256": digest,
        "source_size_bytes": source.stat().st_size,
        "source_sheets": list(rows),
        "selected_public_rows": {
            sheet: sorted(row_numbers)
            for sheet, row_numbers in rows.items()
        },
        "included_fields": ["company", "public_name_or_anonymous_marker", "role_title", "public_source_url", "public_relevance_context"],
        "excluded_private_fields": ["phone", "email", "alternate_email", "telegram", "messenger_report", "outreach_activity", "private_social_profile", "photo_binary"],
        "raw_workbook_copied_to_repository": False,
        "private_contact_values_retained": False,
        "acceptance_basis": "User supplied the workbook and explicitly requested that it be used as the benchmark on 2026-07-17.",
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--accepted-by-user", action="store_true")
    parser.add_argument("--accepted-at", default=datetime.now(UTC).isoformat())
    args = parser.parse_args(argv)
    if not args.accepted_by_user:
        raise SystemExit("Refusing to freeze benchmark without --accepted-by-user")

    rows = _workbook_public_rows(args.source)
    _verify_source_rows(rows)
    benchmark = _benchmark()
    benchmark.assert_no_blind_leakage(benchmark.planning_payload(guided=False))

    args.output_dir.mkdir(parents=True, exist_ok=True)
    benchmark_path = args.output_dir / "benchmark.user.json"
    source_record_path = args.output_dir / "benchmark.source.json"
    freeze_path = args.output_dir / "benchmark.freeze.json"
    benchmark_path.write_text(benchmark.model_dump_json(indent=2) + "\n", encoding="utf-8")
    source_record_path.write_text(
        json.dumps(_source_record(source=args.source, rows=rows), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    freeze = PowerWebBenchmarkFreeze(
        benchmark_path="docs/radar/pipelines/power-web-discovery/benchmark/benchmark.user.json",
        benchmark_sha256=benchmark_sha256(benchmark_path),
        benchmark_id=benchmark.benchmark_id,
        benchmark_version=benchmark.benchmark_version,
        accepted_by_user=True,
        accepted_at=args.accepted_at,
    )
    freeze_path.write_text(freeze.model_dump_json(indent=2) + "\n", encoding="utf-8")
    print(f"benchmark={benchmark.benchmark_id}@{benchmark.benchmark_version}")
    print(f"profiles={len(benchmark.blind_controls.profiles)}")
    print(f"identity_pairs={len(benchmark.blind_controls.identity_pairs)}")
    print(f"employment_controls={len(benchmark.blind_controls.employment)}")
    print(f"relationship_controls={len(benchmark.blind_controls.relationships)}")
    print(f"benchmark_sha256={freeze.benchmark_sha256}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
